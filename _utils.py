import datetime
import glob
import os
import pickle
import random
import re
import subprocess
import threading
import time
from copy import deepcopy
from functools import reduce

import math
import numpy as np
import torch
import torch.nn.functional as F
from loguru import logger
from openpyxl import Workbook  # pip install openpyxl
from openpyxl.reader.excel import load_workbook
from torch import optim
from torch.optim.lr_scheduler import MultiStepLR, CosineAnnealingLR, StepLR, ReduceLROnPlateau
from torchvision import transforms
from tqdm import tqdm

from _data import get_topk, build_loader, get_class_num
import scipy.io as scio


def rename_output(args):
    if os.path.exists(f"./output/{args.backbone}"):
        m_vals = [int(x.split("-m")[1]) for x in os.listdir("./output") if len(x.split("-m")) == 2]
        m_next = (max(m_vals) if len(m_vals) > 0 else 0) + 1
        print(f"rename current output {args.backbone} -> {args.backbone}-m{m_next}")
        os.rename(f"./output/{args.backbone}", f"./output/{args.backbone}-m{m_next}")


def seed_everything(seed=42):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False


def save_checkpoint(args, best_checkpoint):
    if best_checkpoint is None:
        logger.warning(f"no further improvement")
        return
    if "memo" in best_checkpoint:
        memo = best_checkpoint.pop("memo")
        torch.save(best_checkpoint, f"{args.save_dir}/e{best_checkpoint['epoch']}_{memo}.pkl")
    else:
        torch.save(best_checkpoint, f"{args.save_dir}/e{best_checkpoint['epoch']}_{best_checkpoint['map']:.3f}.pkl")


def load_checkpoint(args):
    pkl_list = [x for x in os.listdir(args.save_dir) if x.endswith(".pkl")]
    if len(pkl_list) == 0:
        logger.warning(f"no checkpoint found")
        return None
    pkl_list.sort(key=lambda x: int(x.split("_")[0].replace("e", "")), reverse=True)
    check_point = torch.load(args.save_dir + "/" + pkl_list[0])
    return check_point


def load_selected_states(model, state_dict, startswith_filter=None, verbose=True):
    if startswith_filter is None:
        new_checkpoint = state_dict
    else:
        new_checkpoint = {}
        for k, v in state_dict.items():
            if any(k.startswith(prefix) for prefix in startswith_filter):
                new_checkpoint[k] = v

    missing, unexpected = model.load_state_dict(new_checkpoint, strict=False)
    if unexpected:
        raise RuntimeError("Unexpected state dict keys: {}.".format(unexpected))
    if verbose:
        print(missing if missing else "<All keys matched successfully>")

    return model


def validate_clear():
    global validation_thread
    if "validation_thread" in globals():
        del validation_thread
    # torch.cuda.empty_cache()


def validate_smart(args, query_loader, dbase_loader, early_stopping, epoch, **kwargs):
    global validation_thread

    if "validation_thread" in globals():
        if validation_thread.is_alive():
            print("thread is still running, waiting for it to finish...")
            validation_thread.join()
        if early_stopping.early_stop:
            validate_clear()
            # no need for further validation
            return True

    # for resume training
    kwargs["cur_rng"] = {
        "random_state": random.getstate(),
        "np_random_state": np.random.get_state(),
        "torch_cpu_rng_state": torch.get_rng_state(),
        "torch_cuda_rng_state": torch.cuda.get_rng_state(),
    }

    val_fnc = kwargs.pop("val_fnc", validate)
    parallel_val = kwargs.pop("parallel_val", True)

    if parallel_val and (epoch + 1) != args.n_epochs and early_stopping.counter < early_stopping.patience - 1:
        available_gpu = next((x for x in get_gpu_info() if x["index"] != args.device.split(":")[1]), None)
        if available_gpu:
            print(f'using gpu:{available_gpu["index"]} ({available_gpu["gpu_util"]}%) for evaluation with threading...')

            for k, v in kwargs.items():
                if "model" in k:
                    # validation_model.load_state_dict(kwargs["model"].state_dict())
                    kwargs[k] = deepcopy(v).to(torch.device(f"cuda:{available_gpu['index']}"))
            kwargs["verbose"] = False

            # start validation in a separate thread
            validation_thread = threading.Thread(
                target=val_fnc, args=(args, query_loader, dbase_loader, early_stopping, epoch), kwargs=kwargs
            )
            validation_thread.start()
            return False

    # perform validation without threading
    early_stop = val_fnc(args, query_loader, dbase_loader, early_stopping, epoch, **kwargs)
    if early_stop or (epoch + 1) == args.n_epochs:
        validate_clear()

    return early_stop


def save_mat(query_img, query_labels, retrieval_img, retrieval_labels, args):
    save_dir = './result'
    os.makedirs(save_dir, exist_ok=True)

    query_img = query_img.cpu().detach().numpy()
    retrieval_img = retrieval_img.cpu().detach().numpy()
    query_labels = query_labels.cpu().detach().numpy()
    retrieval_labels = retrieval_labels.cpu().detach().numpy()

    result_dict = {
        'q_img': query_img,
        'r_img': retrieval_img,
        'q_l': query_labels,
        'r_l': retrieval_labels
    }
    scio.savemat(os.path.join(save_dir, args.dataset + str(args.n_bits) + ".mat"), result_dict)


def validate(args, query_loader, dbase_loader, early_stopping, epoch, **kwargs):
    out_idx = kwargs.pop("out_idx", None)
    verbose = kwargs.pop("verbose", True)
    cur_rng = kwargs.pop("cur_rng", None)
    map_fnc = kwargs.pop("map_fnc", mean_average_precision)

    qB, qL = predict(kwargs["model"], query_loader, out_idx=out_idx, verbose=verbose)
    rB, rL = predict(kwargs["model"], dbase_loader, out_idx=out_idx, verbose=verbose)
    map_v = map_fnc(qB, rB, qL, rL, args.topk)

    # save_mat(qB, qL, rB, rL, args)

    map_k = "" if args.topk is None else f"@{args.topk}"

    map_o = early_stopping.best_map
    if map_v > map_o:
        save_mat(qB, qL, rB, rL, args)

    del qB, rB, qL, rL
    torch.cuda.empty_cache()

    early_stopping(epoch, map_v.item(), cur_rng, **kwargs)
    logger.info(
        f"[Evaluating][dataset:{args.dataset}][bits:{args.n_bits}][epoch:{epoch}/{args.n_epochs - 1}][best-mAP{map_k}:{map_o:.4f}][mAP{map_k}:{map_v:.4f}][count:{early_stopping.counter}]"
    )
    return early_stopping.early_stop


class EarlyStopping:
    def __init__(self, patience=10, best_epoch=0, best_map=0.0):
        self.patience = patience
        self.best_epoch = best_epoch
        self.best_map = best_map
        self.best_checkpoint = None
        self.early_stop = False
        self.counter = 0

    def reset(self):
        self.patience = 10
        self.best_epoch = 0
        self.best_map = 0.0
        self.best_checkpoint = None
        self.early_stop = False
        self.counter = 0

    def __call__(self, current_epoch, current_map, current_rng, **kwargs):
        if current_map > self.best_map:
            self.best_epoch, self.best_map = current_epoch, current_map
            self.counter = 0
            self.best_checkpoint = {
                "epoch": self.best_epoch,
                "map": self.best_map,
            }
            memo = kwargs.pop("memo", None)
            if memo:
                self.best_checkpoint["memo"] = memo
            if current_rng:
                self.best_checkpoint.update(current_rng)
            # for k, v in kwargs.items():
            #     self.best_checkpoint[k] = deepcopy(v.state_dict())
            self.best_checkpoint.update({k: deepcopy(v.state_dict()) for k, v in kwargs.items()})
        else:
            self.counter += 1
            if self.counter >= self.patience:
                self.early_stop = True


def print_in_md(rst):
    """
    print training result in Markdown format, like below:
    | K\D |  cifar   | nuswide  |  flickr  |   coco   |
    |----:|:--------:|:--------:|:--------:|:--------:|
    |  16 |    -     | 0.222@09 | 0.333@99 | 0.444@99 |
    |  32 | 0.111@99 | 0.222@99 | 0.333@09 |    -     |
    |  64 | 0.111@99 | 0.222@99 | 0.333@99 | 0.444@99 |
    | 128 |    -     |    -     | 0.333@99 | 0.444@99 |
    """
    if len(rst) == 0:
        return
    datasets = sorted(set([x["dataset"] for x in rst]), key=lambda x: get_class_num(x))
    hash_bits = sorted(set([x["hash_bit"] for x in rst]))
    best_epochs = sorted(set([x["best_epoch"] for x in rst]))

    max_len_e = len(str(best_epochs[-1]))
    max_len_k = max(len(str(hash_bits[-1])), 3)  # len("K\D") -> 3

    head = f"| {'K'+chr(92)+'D':^{max_len_k}} |"
    sept = f"|{':':->{max_len_k+2}}|"

    for dataset in datasets:
        head += f"{dataset:^{max_len_e+8}}|"  # len(" 0.xxx@ ") -> 8
        sept += f":{'-':-^{max_len_e+6}}:|"

    rows = []
    for hash_bit in hash_bits:
        temp_str = f"| {hash_bit:>{max_len_k}} |"
        for dataset in datasets:
            ret = [x for x in rst if x["dataset"] == dataset and x["hash_bit"] == hash_bit]
            if len(ret) == 1:
                temp_str += f" {ret[0]['best_map']:.3f}@{ret[0]['best_epoch']:0>{max_len_e}} |"
            else:
                temp_str += f"{'-':^{max_len_e+8}}|"
        rows.append(temp_str)

    print(head)
    print(sept)
    for row in rows:
        print(row)


def gen_test_data(B, C, K, is_multi_hot=False, normalize_embeddings=True):
    """
    Args:
        B: batch size
        C: number of classes
        K: dim of embeddings
        is_multi_hot: is multi-label dataset or not
        normalize_embeddings: normalize embeddings or not
    Returns:
        embeddings: [B, K]
        singles: [B, ],  categorical ids, None if is_multi_hot
        onehots: [B, C], onehot encoded categorical ids
    """
    embeddings = torch.randn(B, K)
    if is_multi_hot:
        singles = None
        onehots = (torch.randn(B, C) > 0.8).float()
    else:
        singles = torch.randint(low=0, high=C, size=[B])  # categorical id
        onehots = F.one_hot(singles, C).float()
    if normalize_embeddings:
        embeddings = F.normalize(embeddings, p=2, dim=-1)
    return embeddings, singles, onehots


def build_optimizer(optim_type, parameters, **kwargs):
    # optimizer_names = ["Adam", "RMSprop", "SGD"]
    # optimizer = getattr(optim, optimizer_name)(model.parameters(), lr=lr)
    if optim_type == "sgd":
        optimizer = optim.SGD(parameters, **kwargs)
    elif optim_type == "rmsprop":
        optimizer = optim.RMSprop(parameters, **kwargs)
    elif optim_type == "adam":
        optimizer = optim.Adam(parameters, **kwargs)
    elif optim_type == "amsgrad":
        optimizer = optim.Adam(parameters, amsgrad=True, **kwargs)
    elif optim_type == "adamw":
        optimizer = optim.AdamW(parameters, **kwargs)
    else:
        raise NotImplementedError(f"not support optimizer: {optim_type}")
    return optimizer


def build_scheduler(scheduler_type, optimizer, **kwargs):
    if scheduler_type == "none":
        scheduler = None
    elif scheduler_type == "step":
        if "milestones" in kwargs:
            scheduler = MultiStepLR(optimizer, **kwargs)
        else:
            scheduler = StepLR(optimizer, **kwargs)
    elif scheduler_type == "cosine":
        scheduler = CosineAnnealingLR(optimizer, **kwargs)
    elif scheduler_type == "reduce":
        scheduler = ReduceLROnPlateau(optimizer, **kwargs)
    else:
        raise NotImplementedError(f"not support scheduler: {scheduler_type}")
    return scheduler


def get_gpu_info():
    command = [
        "nvidia-smi",
        "--query-gpu=index,name,utilization.gpu,utilization.memory,memory.total,memory.used,memory.free",
        "--format=csv,noheader,nounits",
    ]

    try:
        result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=True)
    except subprocess.CalledProcessError as e:
        print(f"{e.stderr.strip() or e.stdout.strip()}")
        return None
    except FileNotFoundError:
        print("Error: nvidia-smi not found. Ensure NVIDIA drivers and CUDA are installed correctly.")
        return None

    gpu_info = result.stdout.strip().split("\n")

    info_keys = ["index", "name", "gpu_util", "mem_util", "mem_total", "mem_used", "mem_free"]

    return [dict(zip(info_keys, info.split(", "))) for info in gpu_info]


def init():
    assert torch.cuda.is_available(), "CUDA is not available"
    # batch_size is too small will cause "Too many open files..."
    torch.multiprocessing.set_sharing_strategy("file_system")


def find_diff_same(t1, t2, dim):
    """
    Args:
        t1 = torch.tensor([1, 9, 12, 5, 24])
        t2 = torch.tensor([1, 24])
    Returns:
        diff: torch.tensor([5, 9, 12])
        same: torch.tensor([1, 24])
    From:
        https://stackoverflow.com/questions/55110047
    """
    t1 = torch.unique(t1, dim=dim)
    t2 = torch.unique(t2, dim=dim)
    combined = torch.cat((t1, t2))
    uniques, counts = combined.unique(dim=dim, return_counts=True)
    difference = uniques[counts == 1]
    intersection = uniques[counts > 1]
    return difference, intersection


def calc_learnable_params(*args):
    n_parameters = 0
    for x in args:
        if x:
            n_parameters += sum(p.numel() for p in x.parameters() if p.requires_grad)
    return n_parameters


def sizeof_fmt(num, suffix="B"):
    for unit in ("", "Ki", "Mi", "Gi", "Ti", "Pi", "Ei", "Zi"):
        if abs(num) < 1024.0:
            return f"{num:3.1f}{unit}{suffix}"
        num /= 1024.0
    return f"{num:.1f}Yi{suffix}"


def hash_center_type(n_classes, n_bits):
    """
    used in CenterHashing, CSQ, ...
    """
    lg2 = 0 if n_bits < 1 else int(math.log(n_bits, 2))
    if 2**lg2 != n_bits:
        return "random"

    if n_classes <= n_bits:
        return "ha_d"
    elif n_classes > n_bits and n_classes <= 2 * n_bits:
        return "ha_2d"
    else:
        return "random"


def predict(net, dataloader, in_idx=0, out_idx=None, use_sign=True, verbose=True):
    device = next(net.parameters()).device
    # TODO: no need of clses
    codes, clses = [], []
    net.eval()

    # logger.info(f"predicting({len(dataloader.dataset)})...")
    if verbose:
        try:
            _iter = tqdm(dataloader, desc=f"extracting {dataloader.dataset.usage} features")
        except AttributeError:
            _iter = tqdm(dataloader, desc=f"extracting features")
    else:
        _iter = dataloader

    # for img, cls, _ in tqdm(dataloader):
    # for img, tag, cls, _ in tqdm(dataloader): # CrossModal
    for x in _iter:
        with torch.no_grad():
            out = net(x[in_idx].to(device))
        if out_idx is None:
            rst = out
        elif isinstance(out_idx, list):
            rst = reduce(lambda d, key: d[key], out_idx, out)
        else:
            rst = out[out_idx]
        codes.append(rst)
        clses.append(x[-2])
    return torch.cat(codes).sign() if use_sign else torch.cat(codes), torch.cat(clses).to(device)


def mean_average_precision(qB, rB, qL, rL, topk=None):
    """
    Calculate mean average precision(map).

    Args:
        qB (torch.Tensor): Query data hash code.
        rB (torch.Tensor): Database data hash code.
        qL (torch.Tensor): Query data targets, one-hot
        rL (torch.Tensor): Database data targets, one-hot
        topk (Any): Calculate top k data mAP.

    Returns:
        meanAP (float): Mean Average Precision.
    """
    num_query = qL.shape[0]
    if topk is None:
        topk = rL.shape[0]
    mean_AP = 0.0
    for i in range(num_query):
        # Retrieve images from database
        if len(qL.shape) == 1:
            retrieval = (rL == qL[i]).float()
        elif len(qL.shape) == 2:
            retrieval = (qL[i, :] @ rL.T > 0).float()
        else:
            raise NotImplementedError(f"not support: {qL.shape}")
        # Calculate hamming distance
        hamming_dist = 0.5 * (rB.shape[1] - qB[i, :] @ rB.T)
        # Arrange position according to hamming distance
        retrieval = retrieval[torch.argsort(hamming_dist)][:topk]
        # Retrieval count
        retrieval_cnt = retrieval.sum().int().item()
        # Can not retrieve images
        if retrieval_cnt == 0:
            continue
        # Generate score for every position
        score = torch.linspace(1, retrieval_cnt, retrieval_cnt).to(retrieval.device)
        # Acquire index
        index = ((retrieval == 1).nonzero(as_tuple=False).squeeze() + 1.0).float()
        mean_AP += (score / index).mean()
    mean_AP = mean_AP / num_query
    return mean_AP


def calc_hamming_dist(B1, B2):
    """
    calc Hamming distance
    Args:
        B1 (torch.Tensor): each bit of B1 ∈ {-1, 1}^k
        B2 (torch.Tensor): each bit of B2 ∈ {-1, 1}^k
    Returns:
        Hamming distance ∈ {0, k}
    """
    k = B2.shape[1]
    if len(B1.shape) < 2:
        B1 = B1.unsqueeze(0)
    distH = 0.5 * (k - B1.mm(B2.t()))
    return distH


def pr_curve(qB, rB, query_label, retrieval_label):
    num_query = qB.shape[0]
    num_bit = qB.shape[1]
    P = torch.zeros(num_query, num_bit + 1)
    R = torch.zeros(num_query, num_bit + 1)
    for i in range(num_query):
        gnd = (query_label[i].unsqueeze(0).mm(retrieval_label.t()) > 0).float().squeeze()
        tsum = torch.sum(gnd)
        if tsum == 0:
            continue
        hamm = calc_hamming_dist(qB[i, :], rB)
        tmp = (hamm <= torch.arange(0, num_bit + 1).reshape(-1, 1).float().to(hamm.device)).float()
        total = tmp.sum(dim=-1)
        total = total + (total == 0).float() * 0.1
        t = gnd * tmp
        count = t.sum(dim=-1)
        p = count / total
        r = count / tsum
        P[i] = p
        R[i] = r
    mask = (P > 0).float().sum(dim=0)
    mask = mask + (mask == 0).float() * 0.1
    P = P.sum(dim=0) / mask
    R = R.sum(dim=0) / mask
    return P, R


def p_topK(qB, rB, qL, rL, K=None):
    if K is None:
        K = [1, 100, 200, 300, 400, 500, 600, 700, 800, 900, 1000]
    num_query = qL.shape[0]
    p = [0] * len(K)
    for iter in range(num_query):
        gnd = (qL[iter].unsqueeze(0).mm(rL.t()) > 0).float().squeeze()
        tsum = torch.sum(gnd)
        if tsum == 0:
            continue
        hamm = calc_hamming_dist(qB[iter, :], rB).squeeze()
        for i in range(len(K)):
            total = min(K[i], rL.shape[0])
            ind = torch.sort(hamm)[1][:total]
            gnd_ = gnd[ind]
            p[i] += gnd_.sum() / total
    p = torch.Tensor(p) / num_query
    return p


def cos(A, B=None):
    """cosine"""
    # An = normalize(A, norm='l2', axis=1)
    An = A / np.linalg.norm(A, ord=2, axis=1)[:, np.newaxis]
    if (B is None) or (B is A):
        return np.dot(An, An.T)
    # Bn = normalize(B, norm='l2', axis=1)
    Bn = B / np.linalg.norm(B, ord=2, axis=1)[:, np.newaxis]
    return np.dot(An, Bn.T)


def hamming(A, B=None):
    """A, B: [None, bit]
    elements in {-1, 1}
    """
    if B is None:
        B = A
    bit = A.shape[1]
    return (bit - A.dot(B.T)) // 2


def euclidean(A, B=None, sqrt=False):
    aTb = np.dot(A, B.T)
    if (B is None) or (B is A):
        aTa = np.diag(aTb)
        bTb = aTa
    else:
        aTa = np.diag(np.dot(A, A.T))
        bTb = np.diag(np.dot(B, B.T))
    D = aTa[:, np.newaxis] - 2.0 * aTb + bTb[np.newaxis, :]
    if sqrt:
        D = np.sqrt(D)
    return D


def NDCG(qF, rF, qL, rL, what=0, k=-1):
    """Normalized Discounted Cumulative Gain
    ref: https://github.com/kunhe/TALR/blob/master/%2Beval/NDCG.m
    """
    n_query = qF.shape[0]
    if (k < 0) or (k > rF.shape[0]):
        k = rF.shape[0]
    Rel = np.dot(qL, rL.T).astype(int)
    G = 2**Rel - 1
    D = np.log2(2 + np.arange(k))
    if what == 0:
        Rank = np.argsort(1 - cos(qF, rF))
    elif what == 1:
        Rank = np.argsort(hamming(qF, rF))
    elif what == 2:
        Rank = np.argsort(euclidean(qF, rF))

    _NDCG = 0
    for g, rnk in zip(G, Rank):
        dcg_best = (np.sort(g)[::-1][:k] / D).sum()
        if dcg_best > 0:
            dcg = (g[rnk[:k]] / D).sum()
            _NDCG += dcg / dcg_best
    return _NDCG / n_query

class MyEval(object):
    main_dir, proj_name, backbone, dataset, n_bits = None, None, None, None, None

    def __init__(
        self,
        get_config_func,
        build_model_func,
        build_trans_func,
        build_loader_func,
        get_class_num_func,
        get_topk_func,
        proj_order,
        data_order,
        bits_order,
        device,
        the_suffix="",
    ):
        # functions for Supervised or ZeroShot Learning
        self.get_config = get_config_func
        self.build_model = build_model_func
        self.build_trans = build_trans_func
        self.build_loader = build_loader_func
        self.get_class_num = get_class_num_func
        self.get_topk = get_topk_func
        # Variables
        self.proj_order = proj_order
        self.data_order = data_order
        self.bits_order = bits_order
        self.the_suffix = the_suffix
        self.device = device

    def prepare_excel(self, file_path, sheet_name):
        """
        Prepare the excel file to be written, create it if it doesn't exist.
        """
        is_new = False
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            is_new = True
            if "Sheet" in wb.sheetnames:
                ws = wb["Sheet"]
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
        return wb, ws, is_new

    def get_ds_name(self, idx=-1):
        """
        Get the name of the database according to the different purposes.
        For example:
        1)self.dataset = "cifar"
          idx =  *: "cifar"
        2)self.dataset = (nus, voc)
          idx = -1: "nus->voc"
          idx =  0: "nus"
          idx =  1: "voc"
        """
        if isinstance(self.dataset, str):
            return self.dataset
        else:
            if idx == -1:
                return f"{self.dataset[0]}->{self.dataset[1]}"
            return self.dataset[idx]

    def write_excel_map(self, file_path, v):
        """
        support: NDCG, mAP, TrainingTime & EncodingTime
        """
        wb, ws, is_new = self.prepare_excel(file_path, self.proj_name)
        if is_new:
            for i, x in enumerate(self.bits_order["mAP"], 2):
                ws.cell(row=1, column=i).value = f"{x}bits"
            for i, x in enumerate(self.data_order, 2):
                ws.cell(row=i, column=1).value = x

        row1 = [col[0].value for col in ws.iter_cols(min_row=1, max_row=1)]
        col1 = [row[0].value for row in ws.iter_rows(min_col=1, max_col=1)]

        j = row1.index(f"{self.n_bits}bits") + 1
        i = col1.index(self.get_ds_name()) + 1

        ws.cell(row=i, column=j).value = v
        wb.save(file_path)

    def write_excel_hamming2(self, file_path, v):
        """
        save P@H≤2 result to excel.
        """
        wb, ws, is_new = self.prepare_excel(file_path, self.get_ds_name())
        if is_new:
            for i in range(len(self.proj_order)):
                ws.cell(row=1, column=i + 2).value = self.proj_order[i]
            for i, x in enumerate(self.bits_order["P@H≤2"], 2):
                ws.cell(row=i, column=1).value = f"{x}bits"

        row1 = [col[0].value for col in ws.iter_cols(min_row=1, max_row=1)]
        col1 = [row[0].value for row in ws.iter_rows(min_col=1, max_col=1)]

        j = row1.index(self.proj_name) + 1
        i = col1.index(f"{self.n_bits}bits") + 1

        ws.cell(row=i, column=j).value = v
        wb.save(file_path)

    def write_excel_pr(self, file_path, P, R):
        """
        save PR-curve result to excel.
        """
        wb, ws, is_new = self.prepare_excel(file_path, f"{self.get_ds_name()}@{self.n_bits}")
        if is_new:
            for i, x in enumerate(self.proj_order):
                ws.cell(row=1, column=2 * i + 1).value = x
                ws.cell(row=2, column=2 * i + 1).value = "R"
                ws.cell(row=2, column=2 * i + 2).value = "P"

        row1 = [col[0].value for col in ws.iter_cols(min_row=1, max_row=1)]
        try:
            j = row1.index(self.proj_name) + 1
        except ValueError:
            j = len(row1) + 1
            ws.cell(row=1, column=j).value = self.proj_name
            ws.cell(row=2, column=j).value = "R"
            ws.cell(row=2, column=j + 1).value = "P"

        for i, x in enumerate(R):
            ws.cell(row=i + 3, column=j).value = x.item()
            ws.cell(row=i + 3, column=j + 1).value = P[i].item()

        wb.save(file_path)

    def write_excel_topk(self, file_path, rst):
        """
        save TopN-precision result to excel.
        """
        wb, ws, is_new = self.prepare_excel(file_path, f"{self.get_ds_name()}@{self.n_bits}")

        j = 1
        if is_new:
            for i in range(len(self.proj_order)):
                ws.cell(row=1, column=i + 1).value = self.proj_order[i]
        else:
            while True:
                if ws.cell(row=1, column=j).value is None:
                    ws.cell(row=1, column=j).value = self.proj_name
                    break
                if ws.cell(row=1, column=j).value != self.proj_name:
                    j += 1
                    continue
                break

        for i in range(len(rst)):
            ws.cell(row=i + 2, column=j).value = rst[i].item()

        wb.save(file_path)

    def gen_cache_path(self):
        """
        Generate file path of the cache.
        """
        if isinstance(self.dataset, str):
            cache_path = f"{self.main_dir}/{self.proj_name}/output/{self.backbone}/{self.dataset}/{self.n_bits}/cache.p"
        else:
            if self.backbone is None:
                cache_path = (
                    f"{self.main_dir}/{self.proj_name}/output/{self.dataset[0]}_{self.dataset[1]}/{self.n_bits}/cache.p"
                )
            else:
                cache_path = f"{self.main_dir}/{self.proj_name}/output/{self.backbone}/{self.dataset[0]}/{self.n_bits}/{self.dataset[1]}_cache.p"
        return cache_path

    def get_checkpoint_path(self):
        """
        Get the file path of the checkpoint.
        """
        pkl_dir = os.path.dirname(self.gen_cache_path())
        pkl_list = glob.glob(f"{pkl_dir}/*.pkl")
        if len(pkl_list) != 1:
            logger.error(pkl_list)
            raise Exception(f"cannot locate one *.pkl in: {pkl_dir}")
        return pkl_list[0]

    def load_model(self):
        """
        Load pre-trained model based on configurations.
        """
        args = self.get_config()
        if isinstance(args, dict):
            args["backbone"] = self.backbone
            args["device"] = self.device
            args["n_bits"] = self.n_bits
            if "n_classes" in args.keys():
                args["n_classes"] = self.get_class_num(self.get_ds_name(0))
        else:
            args.backbone = self.backbone
            args.device = self.device
            args.n_bits = self.n_bits
            if "n_classes" in args:
                args.n_classes = self.get_class_num(self.get_ds_name(0))
        # build model in cuda
        out = self.build_model(args, pretrained=False)
        if isinstance(out, tuple):
            net, out_idx = out
        else:
            net = out
            out_idx = None
        # load checkpoint
        checkpoint_path = self.get_checkpoint_path()
        checkpoint = torch.load(checkpoint_path, map_location="cpu")
        if "model" in checkpoint.keys():
            checkpoint = checkpoint["model"]
        msg = net.load_state_dict(checkpoint)
        logger.info(f"model loaded: {msg}")

        return net, out_idx

    def build_loaders(self):
        """
        Build query & retrieval data loaders based on configurations.
        """
        args = self.get_config()
        dataset = self.get_ds_name(1)
        data_dir = "../_datasets" + self.the_suffix
        if isinstance(args, dict):
            args["dataset"] = dataset
            args["data_dir"] = data_dir
        else:
            args.dataset = dataset
            args.data_dir = data_dir

        if self.build_trans is not None:
            out = self.build_trans(args)
            if isinstance(out, tuple):
                trans = out[1]
            else:
                trans = out
        else:
            logger.debug("use default transforms")
            trans = transforms.Compose(
                [
                    transforms.Resize(256),
                    transforms.CenterCrop(224),
                    transforms.ToTensor(),
                    transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
                ]
            )
        query_loader = self.build_loader(data_dir, dataset, "query", trans, batch_size=100, num_workers=4)
        dbase_loader = self.build_loader(data_dir, dataset, "dbase", trans, batch_size=100, num_workers=4)

        return query_loader, dbase_loader

    def load_prediction(self):
        """
        Load these four values, with cache use cache, without cache recalculate.
        """
        cache_path = self.gen_cache_path()

        if not os.path.exists(cache_path):
            qB, qL, rB, rL = self.calc_prediction()
            save_obj = {
                "qB": qB.cpu(),
                "qL": qL.cpu(),
                "rB": rB.cpu(),
                "rL": rL.cpu(),
            }
            with open(cache_path, "ab") as f:
                pickle.dump(save_obj, f)
        else:
            logger.debug("load qB, qL, rB, rL from cache")
            with open(cache_path, "rb") as f:
                data = pickle.load(f)
            qB, qL, rB, rL = (
                data["qB"].to(self.device),
                data["qL"].to(self.device),
                data["rB"].to(self.device),
                data["rL"].to(self.device),
            )
        return qB, qL, rB, rL

    def calc_prediction(self):
        """
        A pre-trained model is constructed to compute these 4 values.
        """
        net, out_idx = self.load_model()

        global query_loader
        global dbase_loader

        if "query_loader" in globals():
            if query_loader.dataset.name != self.get_ds_name(1):
                logger.debug(f"{query_loader.dataset.name} != {self.get_ds_name(1)}, loader need to be rebuilt")
                del query_loader
                del dbase_loader
            else:
                logger.debug(f"loader is the same, no need to be rebuilt")

        if "query_loader" not in globals():
            query_loader, dbase_loader = self.build_loaders()

        qB, qL = predict(net, query_loader, out_idx=out_idx)
        rB, rL = predict(net, dbase_loader, out_idx=out_idx)

        return qB, qL, rB, rL

    def get_training_time(self):
        """
        Get the training time of the best map through the log file.
        """
        log_path = os.path.dirname(self.gen_cache_path()) + "/train.log"

        with open(log_path, "r") as f:
            lines = f.read().splitlines()
        ret = []
        epoch = 0
        for line in lines:
            m = re.search(r".+\[Training].+\[epoch:([0-9./]+)]\[time:([0-9.]+)].+", line)
            if m:
                ret.append(m.group(2))
                epoch = int(m.group(1).split("/")[0])
                # print(epoch, m.group(2))
        ret = np.array(ret).astype(float)
        assert epoch != 0, f"can't find training time"
        assert len(ret) == epoch + 1, f"may missing some training time"

        idx = 0
        for line in lines[-5:]:
            # print(line)
            m = re.search(r".+best epoch: (\d+).+", line)
            if m:
                idx = int(m.group(1))
                # print(m.group(1))
                break
        assert idx != 0, "can't find best epoch"

        return ret[: idx + 1].sum()

    def get_encoding_time(self):
        """
        Get the encoding time of the model-generated hash based on the following parameters:
        dataset = "voc"
        batch_size = 100
        usage = "query"
        """
        trans = transforms.Compose(
            [
                transforms.Resize((224, 224)),
                transforms.ToTensor(),
                transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
            ]
        )
        dataloader = self.build_loader(
            "../_datasets_zs", "voc", "query", trans, batch_size=100, shuffle=False, num_workers=4
        )
        net, _ = self.load_model()
        net.eval()
        logger.info(f"predicting({len(dataloader.dataset)})...")
        tic = time.time()
        for x in dataloader:
            with torch.no_grad():
                net(x[0].to(self.device))
        toc = time.time()
        return toc - tic

    def get_run_list(self, evals, n_bits):
        """
        Check whether to run an evaluation based on the number of bits.
        """
        rst = []
        for x in evals:
            if n_bits in self.bits_order[x]:
                rst.append(x)
        return rst

    def __call__(
        self,
        main_dir,
        proj_name,
        backbone,
        evals,
        datasets,
        hash_bits,
        use_cache=True,
        save_xlsx=True,
    ):
        self.main_dir = main_dir
        self.proj_name = proj_name
        self.backbone = backbone
        for dataset in datasets:
            self.dataset = dataset
            logger.info(f"processing dataset: {self.get_ds_name()}")

            for hash_bit in hash_bits:
                self.n_bits = hash_bit
                logger.info(f"processing hash-bit: {hash_bit}")

                run_list = self.get_run_list(evals, hash_bit)

                if len(run_list) == 0:
                    logger.info(f"no eval to run, pass")
                    continue

                if self.get_ds_name(0) == self.get_ds_name(1) or proj_name in ["G_MLZSL", "T_MLZSH"]:
                    is_useful = True
                else:
                    is_useful = False

                if not (
                    (
                        not is_useful
                        and set(run_list) <= {"TrainingTime", "EncodingTime", "Check"}
                        or (is_useful and set(run_list) <= {"TrainingTime", "EncodingTime"})
                    )
                ):
                    if use_cache:
                        qB, qL, rB, rL = self.load_prediction()
                    else:
                        qB, qL, rB, rL = self.calc_prediction()

                # calc mAP
                if "mAP" in run_list:
                    topk = self.get_topk(self.get_ds_name(1))
                    map_k = mean_average_precision(qB, rB, qL, rL, topk)
                    if save_xlsx:
                        self.write_excel_map(f"{main_dir}/eval_map{self.the_suffix}.xlsx", f"{map_k:.3f}")
                    logger.info(f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][mAP@{topk}:{map_k:.3f}]")
                else:
                    map_k = None

                # calc PR curve
                if "PR-curve" in run_list:
                    P, R = pr_curve(qB, rB, qL, rL)
                    if save_xlsx:
                        self.write_excel_pr(f"{main_dir}/eval_pr{self.the_suffix}.xlsx", P, R)
                    logger.info(f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][PR-curve is done]")

                # calc TopN precision
                if "TopN-precision" in run_list:
                    rst = p_topK(qB, rB, qL, rL)
                    if save_xlsx:
                        self.write_excel_topk(f"{main_dir}/eval_topk{self.the_suffix}.xlsx", rst)
                    logger.info(f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][TopN-precision is done]")

                if "NDCG" in run_list or "P@H≤2" in run_list:
                    qB = qB.cpu().numpy()
                    rB = rB.cpu().numpy()
                    qL = qL.cpu().numpy()
                    rL = rL.cpu().numpy()

                # calc NDCG
                if "NDCG" in run_list:
                    ndcg = NDCG(qB, rB, qL, rL, what=1, k=1000)
                    if save_xlsx:
                        self.write_excel_map(f"{main_dir}/eval_ndcg{self.the_suffix}.xlsx", f"{ndcg:.3f}")
                    logger.info(f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][NDCG:{ndcg:.3f}]")

                # calc Precision curves within Hamming Radius 2
                if "P@H≤2" in run_list:
                    prec = get_precision_recall_by_Hamming_Radius(rB, rL, qB, qL)
                    if save_xlsx:
                        self.write_excel_hamming2(f"{main_dir}/eval_hamming2{self.the_suffix}.xlsx", prec)
                    logger.info(f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][P@H≤2:{prec:.3f}]")

                if "TrainingTime" in run_list:
                    rst = self.get_training_time()
                    if save_xlsx:
                        self.write_excel_map(f"{main_dir}/training_time{self.the_suffix}.xlsx", rst)
                    logger.info(
                        f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][TrainingTime:{datetime.timedelta(seconds=rst)}]"
                    )

                if "EncodingTime" in run_list:
                    rst = self.get_encoding_time()
                    if save_xlsx:
                        self.write_excel_map(f"{main_dir}/encoding_time{self.the_suffix}.xlsx", rst)
                    logger.info(
                        f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][EncodingTime:{datetime.timedelta(seconds=rst)}]"
                    )

                if "Check" in run_list:
                    # get mAP from check pint file name
                    pkl_fn = os.path.basename(self.get_checkpoint_path())
                    m = re.search(r"e\d+_([0-9.]+)\.pkl", pkl_fn)
                    if m:
                        map_pkl = float(m.group(1))
                    else:
                        m = re.search(r"iter\d+_([0-9.]+)\.pkl", pkl_fn)
                        if m:
                            map_pkl = float(m.group(1))
                        else:
                            raise Exception(f"can't extract mAP from file name: {pkl_fn}")
                    # get mAP from model prediction
                    temp = None
                    if is_useful:
                        if map_k is None:
                            topk = self.get_topk(self.get_ds_name(1))
                            map_k = mean_average_precision(qB, rB, qL, rL, topk)
                    else:
                        topk = self.get_topk(self.get_ds_name(0))
                        temp = self.dataset
                        self.dataset = (self.dataset[0], self.dataset[0])
                        # logger.debug(self.dataset)
                        if use_cache:
                            qB, qL, rB, rL = self.load_prediction()
                        else:
                            qB, qL, rB, rL = self.calc_prediction()
                        map_k = mean_average_precision(qB, rB, qL, rL, topk)

                    map_k = round(map_k.item(), 3)

                    if map_k != map_pkl:
                        # logger.warning(f"map[{map_k}] != map_pkl[{map_pkl}]")
                        logger.warning(
                            f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][Check:failed][mAP@{topk}:{map_k}][mAP@pkl:{map_pkl}]"
                        )
                    else:
                        # logger.debug(f"map[{map_k}] = map_pkl[{map_pkl}]")
                        logger.info(
                            f"[dataset:{self.get_ds_name()}][bits:{hash_bit}][Check:passed][mAP@{topk}:{map_k}][mAP@pkl:{map_pkl}]"
                        )
                    if temp is not None:
                        self.dataset = temp


def init_my_eval(get_config, build_model, build_trans, device):
    proj_order = [
        "DPSH",
        "DSH",
        "CSQ",
        "OrthoHash",
        "IDHN",
        "HyP2",
        "CenterHashing",
        "SWTH",
        "SPRCH",
    ]
    data_order = ["cifar", "nuswide", "flickr", "coco"]
    bits_order = {
        "mAP": [16, 32, 64, 128],
        "NDCG": [16, 32, 64, 128],
        "PR-curve": [16, 32, 48, 64, 128],
        "TopN-precision": [16, 32, 48, 64, 128],
        "P@H≤2": [16, 32, 48, 64],
        "TrainingTime": [16, 32, 48, 64, 128],
        "EncodingTime": [16, 32, 48, 64, 128],
        "Check": [16, 32, 48, 64, 128],
    }
    my_eval = MyEval(
        get_config,
        build_model,
        build_trans,
        build_loader,
        get_class_num,
        get_topk,
        proj_order,
        data_order,
        bits_order,
        device,
    )
    return my_eval


if __name__ == "__main__":
    t1 = torch.tensor([1, 9, 9, 12, 5, 24])
    t2 = torch.tensor([1, 24])
    d, s = find_diff_same(t1, t2, 0)
    print(d)
    print(s)
    x = hash_center_type(60, 32)
    print(x)
